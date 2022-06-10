from cgi import test
from json import load
from re import X
from traceback import print_tb
from openpyxl import load_workbook
from openpyxl.styles import Alignment




book = load_workbook('silence_labels.xlsx')
sheet  = book.active

sheet_cells=[]


temp = []
sheet_cells=[]
for idx, row in enumerate(sheet.iter_rows()):
    row_cells = []
    for cell in row:
        row_cells.append(cell.value)
        
   # print(row_cells)       
    sheet_cells.append(tuple(row_cells))
    text = row_cells[0].upper()+" ("+row_cells[1]+")\n"+row_cells[2]+","+str(row_cells[3])+"\n"+row_cells[5]+","+row_cells[4]
    
    #sheet.merge_cells(start_row=idx+1, start_column=1, end_row=idx+1, end_column=7)
    targetCell = sheet.cell(row=idx+1,column=1)
    sheet.row_dimensions[idx].height = float(43.75)
    sheet.row_dimensions[idx].width = float(43.75)
    targetCell.value = text
    targetCell.alignment = Alignment(horizontal='center', vertical='center') 
    #temp.append(text)
    print(targetCell.value)
  #  print(text+" ||\n")





book.save('silence_labels.xlsx')
print(" lefutott rendesen")